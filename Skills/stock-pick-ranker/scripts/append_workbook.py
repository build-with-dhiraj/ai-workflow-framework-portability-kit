#!/usr/bin/env python3
"""
Append new stocks to the data sheets and RECOMPUTE the ranking sheets of
Substack_Stock_Picks.xlsx from the ranked universe. Run with /usr/bin/python3 (openpyxl).

This is pragmatic glue: it writes VALUES (header styling is preserved; it does not re-apply
conditional fills / color scales). For full 9-sheet formatting parity, dispatch a builder
subagent with references/workbook-schema.md. Always preserves sheets it doesn't touch.

Usage:
  /usr/bin/python3 append_workbook.py <workbook.xlsx> <ranked.json> [--new "Name A,Name B"]

  <ranked.json>  = output of rank.py over the WHOLE universe ({"stocks":[...]} with rank/master/etc.)
  --new          = comma-separated company names that are newly added (appended to data sheets);
                   omit to only refresh ranking sheets.
"""
import json
import re
import sys

import openpyxl


def norm(s):
    s = re.split(r"[\(/]", str(s))[0]
    s = re.sub(r"\b(ltd|limited|pvt|private)\b", "", s, flags=re.I)
    return re.sub(r"[^a-z0-9]", "", s.lower())


def find_sheet(wb, *prefixes):
    for ws in wb.worksheets:
        for p in prefixes:
            if ws.title.lower().startswith(p.lower()):
                return ws
    return None


def header_row_idx(ws, must_contain):
    """Find the row index whose cells contain a known header token."""
    for r in range(1, min(ws.max_row, 12) + 1):
        vals = [str(c.value or "").lower() for c in ws[r]]
        if any(must_contain.lower() in v for v in vals):
            return r
    return 1


def col_index(ws, hdr_row, *candidates):
    """Find the column whose header matches any candidate substring."""
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(hdr_row, c).value or "").lower()
        for cand in candidates:
            if cand.lower() in v:
                return c
    return None


def rewrite_final_v2(ws, stocks):
    hr = header_row_idx(ws, "Rank")
    cols = {
        "rank": col_index(ws, hr, "rank"),
        "stock": col_index(ws, hr, "stock"),
        "by": col_index(ws, hr, "by", "recommended"),
        "theme": col_index(ws, hr, "theme"),
        "master": col_index(ws, hr, "master"),
        "verdict": col_index(ws, hr, "verdict"),
        "method": col_index(ws, hr, "method"),
        "score": col_index(ws, hr, "val score", "score"),
        "conf": col_index(ws, hr, "confidence"),
        "pe": col_index(ws, hr, "p/e", "pe"),
        "notes": col_index(ws, hr, "notes"),
    }
    # clear existing data rows
    for r in range(ws.max_row, hr, -1):
        ws.delete_rows(r)
    for i, s in enumerate(stocks, 1):
        r = hr + i
        def put(key, val):
            if cols.get(key):
                ws.cell(r, cols[key], val)
        put("rank", s.get("rank"))
        put("stock", s.get("company"))
        put("by", s.get("by"))
        put("theme", s.get("theme"))
        put("master", s.get("master"))
        put("verdict", s.get("verdict"))
        put("method", s.get("valuation_method"))
        put("score", s.get("valuation_score_1to5"))
        put("conf", s.get("valuation_confidence"))
        put("pe", s.get("pe"))
        put("notes", s.get("notes", ""))
    return len(stocks)


def rewrite_master(ws, stocks):
    hr = header_row_idx(ws, "Master Rank")
    fcols = {f: col_index(ws, hr, lbl) for f, lbl in
             [("f1", "moat"), ("f2", "mos"), ("f3", "capeff"), ("f4", "antifrag"),
              ("f5", "asymmetry"), ("f6", "mgmt"), ("f7", "converge")]}
    cols = {
        "rank": col_index(ws, hr, "master rank", "rank"),
        "stock": col_index(ws, hr, "stock"),
        "by": col_index(ws, hr, "by", "recommended"),
        "theme": col_index(ws, hr, "theme"),
        "princ": col_index(ws, hr, "principles"),
        "hot": col_index(ws, hr, "hotness"),
        "master": col_index(ws, hr, "master /", "master/"),
        "conf": col_index(ws, hr, "confidence"),
    }
    for r in range(ws.max_row, hr, -1):
        ws.delete_rows(r)
    for i, s in enumerate(stocks, 1):
        r = hr + i
        if cols.get("rank"): ws.cell(r, cols["rank"], i)
        if cols.get("stock"): ws.cell(r, cols["stock"], s.get("company"))
        if cols.get("by"): ws.cell(r, cols["by"], s.get("by"))
        if cols.get("theme"): ws.cell(r, cols["theme"], s.get("theme"))
        for f, c in fcols.items():
            if c: ws.cell(r, c, s.get(f))
        if cols.get("princ"): ws.cell(r, cols["princ"], s.get("principles_score"))
        if cols.get("hot"): ws.cell(r, cols["hot"], s.get("hotness"))
        if cols.get("master"): ws.cell(r, cols["master"], s.get("master"))
        if cols.get("conf"): ws.cell(r, cols["conf"], s.get("confidence", "High"))
    return len(stocks)


def append_data(ws, records, mapping):
    """Append rows to a data sheet, mapping record fields to columns by header substrings."""
    hr = header_row_idx(ws, next(iter(mapping.values()))[0])
    colmap = {field: col_index(ws, hr, *cands) for field, cands in mapping.items()}
    start = ws.max_row + 1
    for j, rec in enumerate(records):
        for field, c in colmap.items():
            if c and rec.get(field) not in (None, ""):
                ws.cell(start + j, c, rec.get(field))
    return len(records)


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    wb_path, ranked_path = sys.argv[1], sys.argv[2]
    new_names = []
    if "--new" in sys.argv:
        new_names = [n.strip() for n in sys.argv[sys.argv.index("--new") + 1].split(",") if n.strip()]
    stocks = json.load(open(ranked_path)).get("stocks", [])
    wb = openpyxl.load_workbook(wb_path)

    msg = []
    ws = find_sheet(wb, "Final Ranking v2")
    if ws:
        msg.append(f"Final Ranking v2: rewrote {rewrite_final_v2(ws, stocks)} rows")
    ws = find_sheet(wb, "Master Ranking")
    if ws:
        msg.append(f"Master Ranking: rewrote {rewrite_master(ws, stocks)} rows")

    if new_names:
        new_set = {norm(n) for n in new_names}
        new_recs = [s for s in stocks if norm(s.get("company", "")) in new_set]
        ws = find_sheet(wb, "Live Financials")
        if ws:
            n = append_data(ws, new_recs, {
                "company": ["stock"], "ticker": ["ticker"], "price_inr": ["price"],
                "market_cap_cr": ["mkt cap", "market cap"], "pe_ttm": ["p/e"], "pb": ["p/b"],
                "roce_pct": ["roce"], "roe_pct": ["roe"], "debt_to_equity": ["debt/equity"],
                "promoter_holding_pct": ["promoter"], "promoter_pledge_pct": ["pledge"],
                "governance_notes": ["governance"]})
            msg.append(f"Live Financials: appended {n} rows")
        ws = find_sheet(wb, "Stock Picks")
        if ws:
            n = append_data(ws, new_recs, {
                "company": ["company"], "ticker": ["ticker"], "by": ["recommended"],
                "sector": ["sector"], "thesis_crux": ["thesis"], "key_risk": ["key risk"]})
            msg.append(f"Stock Picks: appended {n} rows")

    wb.save(wb_path)
    print("\n".join(msg))
    print(f"saved -> {wb_path} ({len(wb.worksheets)} sheets)")
